from __future__ import annotations

import csv
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO, StringIO
from math import ceil
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook

from ..core import current_profile, require_branch_access, require_permission, sb
from ..xls_biff import read_first_sheet_rows

router = APIRouter(prefix="/api/shortages", tags=["shortages"])

MAX_STOCK_BYTES = 20 * 1024 * 1024
MAX_STOCK_ROWS = 100000


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").casefold().strip()
    value = re.sub(r"[\s\u200e\u200f]+", " ", value)
    return value


def _normalize_code(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _text(value)).strip()
    if re.fullmatch(r"[+-]?\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def _numeric_code(value: str) -> str | None:
    text = _normalize_code(value)
    if not re.fullmatch(r"\d+", text):
        return None
    try:
        return str(int(text))
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise HTTPException(422, "تعذر قراءة ملف CSV الخاص بالمخزون")
    rows: list[list[Any]] = []
    for index, row in enumerate(csv.reader(StringIO(decoded))):
        if index >= MAX_STOCK_ROWS:
            break
        rows.append(list(row))
    return rows


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف Excel الخاص بالمخزون") from exc
    try:
        worksheet = workbook.worksheets[0]
        rows: list[list[Any]] = []
        for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if idx >= MAX_STOCK_ROWS:
                break
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        return rows
    finally:
        workbook.close()


def _rows_from_xls(content: bytes) -> list[list[Any]]:
    try:
        return read_first_sheet_rows(content, max_rows=MAX_STOCK_ROWS, max_cols=96)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف .xls الخاص بالمخزون") from exc


async def _read_stock_report(file: UploadFile) -> list[list[Any]]:
    filename = (file.filename or "").lower()
    if not filename.endswith((".csv", ".xls", ".xlsx")):
        raise HTTPException(422, "ارفع تقرير المخزون بصيغة CSV أو XLS أو XLSX")
    content = await file.read(MAX_STOCK_BYTES + 1)
    if not content:
        raise HTTPException(422, "ملف المخزون فارغ")
    if len(content) > MAX_STOCK_BYTES:
        raise HTTPException(413, "حجم تقرير المخزون أكبر من 20 MB")
    if filename.endswith(".csv"):
        return _rows_from_csv(content)
    if filename.endswith(".xls"):
        return _rows_from_xls(content)
    return _rows_from_xlsx(content)


def _header_key(value: Any) -> str | None:
    text = _normalize_name(_text(value)).replace(".", "").replace(":", "")
    compact = text.replace(" ", "")
    if compact in {"رت", "ر ت".replace(" ", "")}:
        return "item_code"
    if "اسم الصنف" in text:
        return "item_name"
    if "الرقم التجاري" in text or "باركود" in text:
        return "barcode"
    if "الكمية" in text:
        return "quantity"
    if "الصلاحي" in text:
        return "expiry"
    if "التصنيف" in text:
        return "classification"
    if text in {"السعر", "سعر"}:
        return "price"
    if "الأجمالي" in text or "الإجمالي" in text or text == "اجمالي":
        return "total"
    return None


def _embedded_header_record(row: list[Any]) -> dict[str, Any] | None:
    header_positions: dict[str, int] = {}
    for index, value in enumerate(row):
        key = _header_key(value)
        if key and key not in header_positions:
            header_positions[key] = index
    required = {"item_code", "item_name", "quantity"}
    if not required.issubset(header_positions):
        return None
    first_header = min(header_positions.values())
    last_header = max(header_positions.values())
    data_start = last_header + 1
    # The pharmacy export places a complete header block followed immediately by
    # its row values (same relative order) on every CSV row.
    record: dict[str, Any] = {}
    for key, header_index in header_positions.items():
        value_index = data_start + (header_index - first_header)
        if value_index < len(row):
            record[key] = row[value_index]
    if not _normalize_code(record.get("item_code")) and not _text(record.get("item_name")):
        return None
    return record


def _tabular_header_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(rows[:30]):
        mapping: dict[str, int] = {}
        for index, value in enumerate(row):
            key = _header_key(value)
            if key and key not in mapping:
                mapping[key] = index
        if {"item_code", "item_name", "quantity"}.issubset(mapping):
            return row_index, mapping
    return None


def _extract_stock(rows: list[list[Any]]) -> dict[str, Any]:
    if not rows:
        raise HTTPException(422, "تقرير المخزون لا يحتوي على بيانات")

    report_date = None
    source_name = None
    records: list[dict[str, Any]] = []

    # First handle the actual pharmacy CSV layout where the header is repeated
    # inside every row before its data values.
    embedded_count = 0
    for row in rows:
        if report_date is None:
            report_date = next((_parse_date(value) for value in row[:5] if _parse_date(value)), None)
        if source_name is None:
            source_name = next((_text(value) for value in row[:5] if "مخزون" in _text(value)), None)
        record = _embedded_header_record(row)
        if record:
            embedded_count += 1
            records.append(record)

    # Also support a normal CSV/XLS/XLSX table with one header row.
    if embedded_count == 0:
        header = _tabular_header_map(rows)
        if not header:
            raise HTTPException(422, "لم نستطع التعرف على أعمدة كود الصنف واسم الصنف والكمية في تقرير المخزون")
        header_row, mapping = header
        for row in rows[header_row + 1 :]:
            record = {key: row[index] if index < len(row) else None for key, index in mapping.items()}
            if _normalize_code(record.get("item_code")) or _text(record.get("item_name")):
                records.append(record)

    if not records:
        raise HTTPException(422, "لم نجد أصنافًا قابلة للقراءة داخل تقرير المخزون")

    by_code: dict[str, dict[str, Any]] = {}
    nameless: list[dict[str, Any]] = []
    for record in records:
        code = _normalize_code(record.get("item_code"))
        name = _text(record.get("item_name"))
        quantity = _number(record.get("quantity"))
        if quantity is None:
            continue
        clean = {
            "item_code": code,
            "item_name": name,
            "quantity": round(float(quantity), 6),
            "barcode": _text(record.get("barcode")),
            "expiry": _text(record.get("expiry")),
            "classification": _text(record.get("classification")),
            "price": _number(record.get("price")),
            "total": _number(record.get("total")),
        }
        if code:
            if code in by_code:
                current = by_code[code]
                current["quantity"] = round(float(current["quantity"]) + float(clean["quantity"]), 6)
                if len(name) > len(str(current.get("item_name") or "")):
                    current["item_name"] = name
                if not current.get("barcode") and clean.get("barcode"):
                    current["barcode"] = clean["barcode"]
            else:
                by_code[code] = clean
        else:
            nameless.append(clean)

    stock_rows = list(by_code.values()) + nameless
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    numeric_codes: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in stock_rows:
        normalized_name = _normalize_name(str(record.get("item_name") or ""))
        if normalized_name:
            by_name[normalized_name].append(record)
        numeric = _numeric_code(str(record.get("item_code") or ""))
        if numeric:
            numeric_codes[numeric].append(record)

    return {
        "report_date": report_date.isoformat() if report_date else None,
        "source_name": source_name,
        "rows": stock_rows,
        "by_code": by_code,
        "by_name": by_name,
        "numeric_codes": numeric_codes,
    }


async def _movement_report(report_id: str, profile: dict[str, Any]) -> dict[str, Any]:
    rows = await sb(
        "GET",
        "/rest/v1/item_movement_reports",
        service=True,
        params={
            "select": "id,branch_id,source_name,source_filename,period_start,period_end,days_count,unique_item_count,unresolved_count,branches(name)",
            "id": f"eq.{report_id}",
            "limit": "1",
        },
    )
    if not rows:
        raise HTTPException(404, "تقرير الحركة غير موجود")
    report = rows[0]
    require_branch_access(profile, str(report["branch_id"]))
    return report


async def _movement_rows(report_id: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(0, 20000, 1000):
        batch = await sb(
            "GET",
            "/rest/v1/item_movement_rows",
            service=True,
            params={
                "select": "id,report_name,item_id,boxes_sold,loose_sold,units_per_box,equivalent_boxes,daily_rate,matched_by,item_catalog(item_code,item_name,package_form,units_per_box)",
                "report_id": f"eq.{report_id}",
                "order": "daily_rate.desc.nullslast,report_name.asc",
            },
            headers={"Range": f"{offset}-{offset + 999}"},
        )
        rows.extend(batch or [])
        if len(batch or []) < 1000:
            break
    return rows


def _match_stock(
    movement: dict[str, Any],
    stock: dict[str, Any],
) -> tuple[dict[str, Any] | None, str]:
    catalog = movement.get("item_catalog") or {}
    code = _normalize_code(catalog.get("item_code"))
    if code and code in stock["by_code"]:
        return stock["by_code"][code], "code"

    numeric = _numeric_code(code) if code else None
    if numeric:
        candidates = stock["numeric_codes"].get(numeric) or []
        if len(candidates) == 1:
            return candidates[0], "code_normalized"

    for name in (catalog.get("item_name"), movement.get("report_name")):
        normalized = _normalize_name(_text(name))
        if not normalized:
            continue
        candidates = stock["by_name"].get(normalized) or []
        if len(candidates) == 1:
            return candidates[0], "name"
    return None, "unmatched"


def _priority(stock_qty: float, days_cover: float, target_days: int, suggested: int) -> tuple[str, str, int]:
    if suggested <= 0:
        return "sufficient", "كافي", 50
    if stock_qty <= 0:
        return "out", "نفد", 0
    if days_cover <= 7:
        return "urgent", "عاجل", 10
    if days_cover <= 14:
        return "soon", "قريب ينقص", 20
    if days_cover < target_days:
        return "monitor", "يحتاج استكمال", 30
    return "sufficient", "كافي", 50


def _analyze_shortages(
    movement_rows: list[dict[str, Any]],
    stock: dict[str, Any],
    target_days: int,
) -> dict[str, Any]:
    result_rows: list[dict[str, Any]] = []
    blocked_rate_count = 0

    for movement in movement_rows:
        daily_rate_raw = _number(movement.get("daily_rate"))
        if daily_rate_raw is None or daily_rate_raw <= 0:
            blocked_rate_count += 1
            continue
        daily_rate = float(daily_rate_raw)
        stock_row, matched_by = _match_stock(movement, stock)
        catalog = movement.get("item_catalog") or {}
        item_code = _normalize_code(catalog.get("item_code"))
        item_name = _text(catalog.get("item_name")) or _text(movement.get("report_name"))

        if stock_row is None:
            result_rows.append({
                "movement_row_id": str(movement.get("id") or ""),
                "item_id": movement.get("item_id"),
                "item_code": item_code or None,
                "item_name": item_name,
                "report_name": _text(movement.get("report_name")),
                "stock_name": None,
                "stock_quantity": None,
                "daily_rate": round(daily_rate, 6),
                "days_cover": None,
                "target_quantity": round(daily_rate * target_days, 3),
                "suggested_quantity": None,
                "status": "unmatched",
                "status_label": "غير موجود في المخزون",
                "priority_rank": 40,
                "matched_by": "unmatched",
                "units_per_box": movement.get("units_per_box") or catalog.get("units_per_box"),
            })
            continue

        stock_qty = max(0.0, float(stock_row.get("quantity") or 0))
        target_quantity = daily_rate * target_days
        shortage_raw = max(0.0, target_quantity - stock_qty)
        suggested = int(ceil(shortage_raw - 1e-9)) if shortage_raw > 0 else 0
        days_cover = stock_qty / daily_rate if daily_rate > 0 else 0.0
        status, status_label, rank = _priority(stock_qty, days_cover, target_days, suggested)
        result_rows.append({
            "movement_row_id": str(movement.get("id") or ""),
            "item_id": movement.get("item_id"),
            "item_code": item_code or _normalize_code(stock_row.get("item_code")) or None,
            "item_name": item_name or _text(stock_row.get("item_name")),
            "report_name": _text(movement.get("report_name")),
            "stock_name": _text(stock_row.get("item_name")),
            "stock_quantity": round(stock_qty, 6),
            "daily_rate": round(daily_rate, 6),
            "days_cover": round(days_cover, 3),
            "target_quantity": round(target_quantity, 3),
            "suggested_quantity": suggested,
            "status": status,
            "status_label": status_label,
            "priority_rank": rank,
            "matched_by": matched_by,
            "units_per_box": movement.get("units_per_box") or catalog.get("units_per_box"),
            "barcode": stock_row.get("barcode") or None,
            "expiry": stock_row.get("expiry") or None,
            "classification": stock_row.get("classification") or None,
        })

    result_rows.sort(
        key=lambda row: (
            int(row.get("priority_rank") or 99),
            float(row.get("days_cover")) if row.get("days_cover") is not None else 10**9,
            -float(row.get("daily_rate") or 0),
            str(row.get("item_name") or ""),
        )
    )

    matched = [row for row in result_rows if row["status"] != "unmatched"]
    shortage_rows = [row for row in matched if int(row.get("suggested_quantity") or 0) > 0]
    summary = {
        "movement_rows": len(movement_rows),
        "rate_ready_rows": len(result_rows),
        "blocked_rate_count": blocked_rate_count,
        "matched_stock_count": len(matched),
        "unmatched_stock_count": sum(1 for row in result_rows if row["status"] == "unmatched"),
        "shortage_count": len(shortage_rows),
        "out_of_stock_count": sum(1 for row in result_rows if row["status"] == "out"),
        "urgent_count": sum(1 for row in result_rows if row["status"] == "urgent"),
        "soon_count": sum(1 for row in result_rows if row["status"] == "soon"),
        "monitor_count": sum(1 for row in result_rows if row["status"] == "monitor"),
        "sufficient_count": sum(1 for row in result_rows if row["status"] == "sufficient"),
        "total_suggested_boxes": sum(int(row.get("suggested_quantity") or 0) for row in shortage_rows),
    }
    return {"rows": result_rows, "summary": summary}


@router.post("/analyze")
async def analyze_shortages(
    movement_report_id: str = Form(...),
    target_days: int = Form(14),
    file: UploadFile = File(...),
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "view_item_analysis")
    if target_days < 1 or target_days > 180:
        raise HTTPException(422, "مدة التغطية يجب أن تكون بين يوم واحد و180 يوم")

    report = await _movement_report(movement_report_id, profile)
    raw_stock_rows = await _read_stock_report(file)
    stock = _extract_stock(raw_stock_rows)
    movements = await _movement_rows(movement_report_id)
    analysis = _analyze_shortages(movements, stock, target_days)

    return {
        "movement_report": report,
        "stock_report": {
            "source_filename": (file.filename or "")[:240],
            "report_date": stock.get("report_date"),
            "source_name": stock.get("source_name"),
            "stock_item_count": len(stock["rows"]),
        },
        "target_days": target_days,
        **analysis,
    }
