from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from ..core import (
    apply_branch_filter,
    current_profile,
    require_branch_access,
    require_permission,
    sb,
)
from ..xls_biff import read_first_sheet_rows

router = APIRouter(prefix="/api/item-movements", tags=["item-movements"])

MAX_REPORT_BYTES = 20 * 1024 * 1024
MAX_REPORT_ROWS = 50000
BOX_UNIT = "علبة"
LOOSE_UNIT = "فرط"


class MovementMapInput(BaseModel):
    item_id: str


class MovementCatalogCreateInput(BaseModel):
    item_code: str | None = Field(default=None, max_length=160)
    item_name: str | None = Field(default=None, max_length=240)
    package_form: str | None = Field(default=None, max_length=120)
    units_per_box: int = Field(gt=0, le=100000)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").casefold().strip()
    return " ".join(value.split())


def _normalize_code(value: Any) -> str:
    text = _text(value)
    if not text or text == "0":
        return ""
    # Excel/CSV exports sometimes render a numeric code as 12345.0.
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return unicodedata.normalize("NFKC", text).strip().casefold()


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    number = _number(value)
    if number is None or number < 1:
        return None
    try:
        return (datetime(1899, 12, 30) + timedelta(days=number)).date()
    except (OverflowError, ValueError):
        return None


def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف Excel") from exc
    try:
        worksheet = workbook.worksheets[0]
        rows: list[list[Any]] = []
        for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
            if idx >= MAX_REPORT_ROWS:
                break
            values = list(row)
            while values and values[-1] is None:
                values.pop()
            rows.append(values)
        return rows
    finally:
        workbook.close()


def _rows_from_xls(content: bytes) -> list[list[Any]]:
    try:
        return read_first_sheet_rows(content, max_rows=MAX_REPORT_ROWS, max_cols=64)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف .xls الصادر من منظومة البيع") from exc


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(422, "تعذر قراءة ترميز ملف CSV")
    try:
        rows: list[list[Any]] = []
        for idx, row in enumerate(csv.reader(io.StringIO(text))):
            if idx >= MAX_REPORT_ROWS:
                break
            rows.append(list(row))
        return rows
    except csv.Error as exc:
        raise HTTPException(422, "تعذر قراءة تقرير CSV") from exc


async def _read_report(file: UploadFile) -> tuple[bytes, list[list[Any]]]:
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xls") or filename.endswith(".xlsx")):
        raise HTTPException(422, "ارفع تقرير المبيعات والحركة بصيغة CSV (أو Excel القديم)")
    content = await file.read(MAX_REPORT_BYTES + 1)
    if not content:
        raise HTTPException(422, "ملف التقرير فارغ")
    if len(content) > MAX_REPORT_BYTES:
        raise HTTPException(413, "حجم تقرير الحركة أكبر من 20 MB")
    if filename.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif filename.endswith(".xls"):
        rows = _rows_from_xls(content)
    else:
        rows = _rows_from_xlsx(content)
    return content, rows


def _value_date(value: Any) -> date | None:
    parsed = _excel_date(value)
    if parsed:
        return parsed
    text = _text(value).split()[0] if _text(value) else ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _label_value(row: list[Any], label: str) -> str:
    target = _normalize_name(label.replace("：", ":"))
    for index, value in enumerate(row):
        current = _normalize_name(_text(value).replace("：", ":"))
        if current == target and index > 0:
            return _text(row[index - 1])
    return ""


def _detail_item_values(row: list[Any]) -> dict[str, str]:
    headers = ["الإجمالي", "السعر", "التعبئة", "الكمية", "اسم الصنف", "ر.ت"]
    positions: list[int] = []
    for header in headers:
        try:
            positions.append(next(i for i, value in enumerate(row) if _text(value) == header))
        except StopIteration:
            return {}
    if positions != list(range(positions[0], positions[0] + len(headers))):
        return {}
    values_start = positions[-1] + 1
    if len(row) < values_start + len(headers):
        return {}
    values = row[values_start:values_start + len(headers)]
    return dict(zip(headers, map(_text, values)))


def _detect_period(rows: list[list[Any]]) -> tuple[date, date, str | None]:
    if not rows:
        raise HTTPException(422, "تقرير الحركة لا يحتوي على بيانات")
    first_rows = rows[:8]
    source_name = None
    if first_rows and first_rows[0]:
        first = first_rows[0]
        is_detailed_csv = any(_normalize_name(_text(x)) in {_normalize_name(": من"), _normalize_name(": إلي"), _normalize_name(": إلى")} for x in first)
        if is_detailed_csv and len(first) > 2:
            # The detailed CSV used by doctor sales stores the pharmacy/source in column 3.
            source_name = _text(first[2]) or None
        if not source_name:
            source_name = next((_text(x) for x in first if _text(x)), None)

    # Detailed CSV layout used by doctor sales: date immediately before : من / : إلي.
    for row in first_rows:
        start_text = _label_value(row, ": من")
        end_text = _label_value(row, ": إلي") or _label_value(row, ": إلى")
        start = _value_date(start_text)
        end = _value_date(end_text)
        if start and end and end >= start:
            return start, end, source_name

    # Preferred legacy Excel layout used by the pharmacy sales system:
    # ... [end date, ': إلي', start date, ': تقرير حركة خلال الفترة من']
    for row in first_rows:
        for index, value in enumerate(row):
            label = _text(value)
            if "تقرير حركة خلال الفترة من" in label:
                start = _value_date(row[index - 1]) if index >= 1 else None
                end = None
                for left in range(index - 2, max(-1, index - 7), -1):
                    if left < 0:
                        break
                    if "إلي" in _text(row[left]) or "إلى" in _text(row[left]):
                        if left >= 1:
                            end = _value_date(row[left - 1])
                        break
                if start and end and end >= start:
                    return start, end, source_name

    dates: list[date] = []
    for row in first_rows:
        for value in row:
            parsed = _value_date(value)
            if parsed and date(2020, 1, 1) <= parsed <= date(2100, 12, 31):
                dates.append(parsed)
    if len(dates) >= 2:
        return min(dates), max(dates), source_name
    raise HTTPException(422, "لم نستطع قراءة فترة التقرير تلقائيًا من ملف حركة الأصناف")


def _aggregate_sales(rows: list[list[Any]]) -> tuple[dict[str, dict[str, Any]], int]:
    """Aggregate net item movement.

    V36 prefers the detailed CSV also used by doctor sales. Cash + credit sales
    increase demand; returns decrease it. Legacy XLS/XLSX movement reports stay
    supported as a fallback.
    """
    aggregates: dict[str, dict[str, Any]] = {}
    transaction_count = 0
    detailed_rows_found = 0

    for row in rows:
        movement_type = _label_value(row, ": نوع الحركة")
        item_values = _detail_item_values(row)
        if not movement_type or not item_values:
            continue

        is_return = "مردود" in movement_type or "مرتجع" in movement_type
        is_sale = movement_type.startswith("مبيعات") and not is_return
        if not (is_sale or is_return):
            continue

        item_name = _text(item_values.get("اسم الصنف"))
        item_code = _text(item_values.get("ر.ت"))
        quantity = _number(item_values.get("الكمية"))
        unit = _text(item_values.get("التعبئة"))
        if not item_name or quantity is None or quantity < 0 or unit not in {BOX_UNIT, LOOSE_UNIT}:
            continue

        detailed_rows_found += 1
        normalized_name = _normalize_name(item_name)
        normalized_code = _normalize_code(item_code)
        aggregate_key = f"code:{normalized_code}" if normalized_code else f"name:{normalized_name}"
        storage_norm = f"code:{normalized_code}|{normalized_name}" if normalized_code else normalized_name
        bucket = aggregates.setdefault(aggregate_key, {
            "report_name": item_name,
            "report_name_norm": storage_norm,
            "match_name_norm": normalized_name,
            "report_code": item_code if normalized_code else "",
            "boxes_sold": 0.0,
            "loose_sold": 0.0,
        })
        if len(item_name) > len(bucket["report_name"]):
            bucket["report_name"] = item_name
            bucket["match_name_norm"] = normalized_name
            bucket["report_name_norm"] = f"code:{normalized_code}|{normalized_name}" if normalized_code else normalized_name
        if not bucket.get("report_code") and normalized_code:
            bucket["report_code"] = item_code

        signed_quantity = -quantity if is_return else quantity
        if unit == BOX_UNIT:
            bucket["boxes_sold"] += signed_quantity
        else:
            bucket["loose_sold"] += signed_quantity
        if is_sale:
            transaction_count += 1

    if detailed_rows_found:
        # A return-heavy period can be net negative. For demand forecasting we do
        # not create a negative sales rate; zero means no positive net demand.
        for bucket in aggregates.values():
            bucket["_signed_boxes"] = round(float(bucket["boxes_sold"]), 6)
            bucket["_signed_loose"] = round(float(bucket["loose_sold"]), 6)
            bucket["boxes_sold"] = max(0.0, bucket["_signed_boxes"])
            bucket["loose_sold"] = max(0.0, bucket["_signed_loose"])
        if not aggregates:
            raise HTTPException(422, "لم نجد حركات مبيعات أو مردودات صالحة داخل تقرير CSV")
        return aggregates, transaction_count

    # Legacy Excel movement report parser.
    for row in rows:
        movement_index = next((i for i, value in enumerate(row) if i >= 3 and _text(value).startswith("مبيعات") and _text(row[i - 3]) in {BOX_UNIT, LOOSE_UNIT}), None)
        if movement_index is None:
            continue

        item_name = _text(row[movement_index - 1])
        quantity = _number(row[movement_index - 2])
        unit = _text(row[movement_index - 3])
        movement_type = _text(row[movement_index])
        if not item_name or quantity is None or quantity < 0 or unit not in {BOX_UNIT, LOOSE_UNIT}:
            continue
        if "مرتجع" in movement_type or "مردود" in movement_type:
            continue

        normalized = _normalize_name(item_name)
        if not normalized:
            continue
        bucket = aggregates.setdefault(f"name:{normalized}", {
            "report_name": item_name,
            "report_name_norm": normalized,
            "match_name_norm": normalized,
            "report_code": "",
            "boxes_sold": 0.0,
            "loose_sold": 0.0,
        })
        if len(item_name) > len(bucket["report_name"]):
            bucket["report_name"] = item_name
        if unit == BOX_UNIT:
            bucket["boxes_sold"] += quantity
        else:
            bucket["loose_sold"] += quantity
        transaction_count += 1

    if not aggregates:
        raise HTTPException(422, "لم نجد حركات مبيعات علبة/فرط داخل التقرير")
    return aggregates, transaction_count


async def _catalog_and_aliases() -> tuple[list[dict[str, Any]], dict[str, str]]:
    catalog: list[dict[str, Any]] = []
    for offset in range(0, 30000, 1000):
        rows = await sb(
            "GET", "/rest/v1/item_catalog", service=True,
            params={"select": "id,item_code,item_name,units_per_box", "order": "item_name.asc"},
            headers={"Range": f"{offset}-{offset + 999}"},
        )
        catalog.extend(rows or [])
        if len(rows or []) < 1000:
            break
    aliases = await sb(
        "GET", "/rest/v1/item_name_aliases", service=True,
        params={"select": "report_name_norm,item_id", "limit": "30000"},
    )
    return catalog, {str(x.get("report_name_norm") or ""): str(x.get("item_id") or "") for x in aliases or []}


def _resolve_rows(
    aggregates: dict[str, dict[str, Any]], catalog: list[dict[str, Any]], aliases: dict[str, str], days_count: int
) -> list[dict[str, Any]]:
    by_id = {str(x["id"]): x for x in catalog}
    by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in catalog:
        code = _normalize_code(item.get("item_code"))
        name = _normalize_name(str(item.get("item_name") or ""))
        if code:
            by_code[code].append(item)
        if name:
            by_name[name].append(item)

    resolved: list[dict[str, Any]] = []
    for _aggregate_key, aggregate in aggregates.items():
        item = None
        matched_by = "unmatched"
        normalized_name = str(aggregate.get("match_name_norm") or aggregate.get("report_name_norm") or "")
        normalized_code = _normalize_code(aggregate.get("report_code"))

        # V36: the detailed CSV carries the real item code. It is authoritative
        # and is tried before the name/learned alias, so renamed items stay one item.
        code_matches = by_code.get(normalized_code) or [] if normalized_code else []
        if len(code_matches) == 1:
            item = code_matches[0]
            matched_by = "exact"
        else:
            exact = by_name.get(normalized_name) or []
            if len(exact) == 1:
                item = exact[0]
                matched_by = "exact"
            else:
                alias_item_id = aliases.get(normalized_name)
                if alias_item_id and alias_item_id in by_id:
                    item = by_id[alias_item_id]
                    matched_by = "alias"

        boxes = round(float(aggregate["boxes_sold"]), 6)
        loose = round(float(aggregate["loose_sold"]), 6)
        signed_boxes = round(float(aggregate.get("_signed_boxes", boxes)), 6)
        signed_loose = round(float(aggregate.get("_signed_loose", loose)), 6)
        units = int(item.get("units_per_box") or 0) if item else None
        equivalent = None
        daily = None
        if item and units and units > 0:
            equivalent = round(max(0.0, signed_boxes + signed_loose / units), 6)
            daily = round(equivalent / days_count, 6)
        elif signed_loose == 0:
            equivalent = max(0.0, signed_boxes)
            daily = round(equivalent / days_count, 6)

        resolved.append({
            **aggregate,
            "item_id": str(item["id"]) if item else None,
            "item_code": str(item.get("item_code") or "") if item else None,
            "catalog_name": str(item.get("item_name") or "") if item else None,
            "units_per_box": units,
            "equivalent_boxes": equivalent,
            "daily_rate": daily,
            "matched_by": matched_by,
        })
    return resolved


async def _parse_and_resolve(file: UploadFile) -> dict[str, Any]:
    _content, rows = await _read_report(file)
    start, end, source_name = _detect_period(rows)
    days_count = (end - start).days + 1
    if days_count <= 0 or days_count > 370:
        raise HTTPException(422, "فترة تقرير الحركة غير صالحة")
    aggregates, transaction_count = _aggregate_sales(rows)
    catalog, aliases = await _catalog_and_aliases()
    resolved = _resolve_rows(aggregates, catalog, aliases, days_count)
    return {
        "period_start": start.isoformat(),
        "period_end": end.isoformat(),
        "days_count": days_count,
        "source_name": source_name,
        "transaction_count": transaction_count,
        "unique_item_count": len(resolved),
        "unresolved_count": sum(1 for x in resolved if not x["item_id"]),
        "blocking_count": sum(1 for x in resolved if not x["item_id"] and float(x["loose_sold"]) > 0),
        "rows": resolved,
    }


async def _report(report_id: str, profile: dict[str, Any]) -> dict[str, Any]:
    rows = await sb(
        "GET", "/rest/v1/item_movement_reports", service=True,
        params={"select": "id,branch_id,period_start,period_end,days_count", "id": f"eq.{report_id}", "limit": "1"},
    )
    if not rows:
        raise HTTPException(404, "تقرير الحركة غير موجود")
    require_branch_access(profile, str(rows[0]["branch_id"]))
    return rows[0]


async def _link_movement_row(
    movement: dict[str, Any],
    report: dict[str, Any],
    item: dict[str, Any],
    profile: dict[str, Any],
) -> dict[str, Any]:
    units = int(item["units_per_box"])
    equivalent = round(
        float(movement.get("boxes_sold") or 0)
        + float(movement.get("loose_sold") or 0) / units,
        6,
    )
    daily = round(equivalent / int(report["days_count"]), 6)
    updated = await sb(
        "PATCH", "/rest/v1/item_movement_rows", service=True,
        headers={"Prefer": "return=representation"}, params={"id": f"eq.{movement['id']}"},
        json={
            "item_id": item["id"], "units_per_box": units,
            "equivalent_boxes": equivalent, "daily_rate": daily, "matched_by": "manual",
        },
    )
    await sb(
        "POST", "/rest/v1/item_name_aliases?on_conflict=report_name_norm", service=True,
        headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
        json={
            "report_name": movement["report_name"], "report_name_norm": _normalize_name(str(movement.get("report_name") or "")),
            "item_id": item["id"], "created_by": profile["id"], "updated_at": datetime.utcnow().isoformat() + "Z",
        },
    )
    unresolved = await sb(
        "GET", "/rest/v1/item_movement_rows", service=True,
        params={"select": "id", "report_id": f"eq.{movement['report_id']}", "item_id": "is.null", "limit": "10000"},
    )
    await sb(
        "PATCH", "/rest/v1/item_movement_reports", service=True,
        params={"id": f"eq.{movement['report_id']}"}, json={"unresolved_count": len(unresolved or [])},
    )
    return updated[0]


@router.post("/preview")
async def preview_report(
    branch_id: str = Form(...),
    file: UploadFile = File(...),
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "view_item_analysis")
    require_branch_access(profile, branch_id)
    parsed = await _parse_and_resolve(file)
    # Preview is intentionally compact: full rows are saved/loaded after import.
    sample_unmatched = [
        {"report_name": x["report_name"], "boxes_sold": x["boxes_sold"], "loose_sold": x["loose_sold"]}
        for x in parsed["rows"] if not x["item_id"]
    ][:20]
    return {key: value for key, value in parsed.items() if key != "rows"} | {"unmatched_sample": sample_unmatched}


@router.post("/import")
async def import_report(
    branch_id: str = Form(...),
    file: UploadFile = File(...),
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    require_branch_access(profile, branch_id)
    parsed = await _parse_and_resolve(file)

    # Re-uploading the exact same branch + period replaces that analysis. Name
    # mappings live in a separate table, so already learned aliases are retained.
    existing = await sb(
        "GET", "/rest/v1/item_movement_reports", service=True,
        params={
            "select": "id", "branch_id": f"eq.{branch_id}",
            "period_start": f"eq.{parsed['period_start']}", "period_end": f"eq.{parsed['period_end']}", "limit": "1",
        },
    )
    if existing:
        await sb("DELETE", "/rest/v1/item_movement_reports", service=True, params={"id": f"eq.{existing[0]['id']}"})

    report_rows = await sb(
        "POST", "/rest/v1/item_movement_reports", service=True,
        headers={"Prefer": "return=representation"},
        json={
            "branch_id": branch_id,
            "source_name": parsed["source_name"],
            "source_filename": (file.filename or "")[:240],
            "period_start": parsed["period_start"],
            "period_end": parsed["period_end"],
            "days_count": parsed["days_count"],
            "transaction_count": parsed["transaction_count"],
            "unique_item_count": parsed["unique_item_count"],
            "unresolved_count": parsed["unresolved_count"],
            "created_by": profile["id"],
        },
    )
    report = report_rows[0]
    payload = []
    for row in parsed["rows"]:
        payload.append({
            "report_id": report["id"],
            "report_name": row["report_name"][:300],
            "report_name_norm": row["report_name_norm"][:300],
            "item_id": row["item_id"],
            "boxes_sold": row["boxes_sold"],
            "loose_sold": row["loose_sold"],
            "units_per_box": row["units_per_box"],
            "equivalent_boxes": row["equivalent_boxes"],
            "daily_rate": row["daily_rate"],
            "matched_by": row["matched_by"],
        })
    for start in range(0, len(payload), 400):
        await sb(
            "POST", "/rest/v1/item_movement_rows", service=True,
            headers={"Prefer": "return=minimal"}, json=payload[start:start + 400],
        )
    return {"ok": True, "report_id": report["id"], **{k: v for k, v in parsed.items() if k != "rows"}}


@router.get("/reports")
async def list_reports(
    branch_id: str | None = None,
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "view_item_analysis")
    params: dict[str, str] = {
        "select": "id,branch_id,source_name,source_filename,period_start,period_end,days_count,transaction_count,unique_item_count,unresolved_count,created_at,branches(name)",
        "order": "period_end.desc,created_at.desc", "limit": "100",
    }
    params = apply_branch_filter(params, profile)
    if branch_id:
        require_branch_access(profile, branch_id)
        params["branch_id"] = f"eq.{branch_id}"
    return await sb("GET", "/rest/v1/item_movement_reports", service=True, params=params)


@router.get("/reports/{report_id}")
async def report_detail(report_id: str, profile: dict[str, Any] = Depends(current_profile)) -> Any:
    require_permission(profile, "view_item_analysis")
    report = await _report(report_id, profile)
    report_full = await sb(
        "GET", "/rest/v1/item_movement_reports", service=True,
        params={
            "select": "id,branch_id,source_name,source_filename,period_start,period_end,days_count,transaction_count,unique_item_count,unresolved_count,created_at,branches(name)",
            "id": f"eq.{report_id}", "limit": "1",
        },
    )
    rows: list[dict[str, Any]] = []
    for offset in range(0, 10000, 1000):
        batch = await sb(
            "GET", "/rest/v1/item_movement_rows", service=True,
            params={
                "select": "id,report_name,item_id,boxes_sold,loose_sold,units_per_box,equivalent_boxes,daily_rate,matched_by,item_catalog(item_code,item_name)",
                "report_id": f"eq.{report_id}", "order": "daily_rate.desc.nullslast,report_name.asc",
            },
            headers={"Range": f"{offset}-{offset + 999}"},
        )
        rows.extend(batch or [])
        if len(batch or []) < 1000:
            break
    total_equivalent = round(sum(float(x.get("equivalent_boxes") or 0) for x in rows), 4)
    blocking = sum(1 for x in rows if not x.get("item_id") and float(x.get("loose_sold") or 0) > 0)
    return {"report": report_full[0] if report_full else report, "rows": rows, "total_equivalent_boxes": total_equivalent, "blocking_count": blocking}


@router.post("/rows/{row_id}/map")
async def map_report_row(
    row_id: str,
    data: MovementMapInput,
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    movement_rows = await sb(
        "GET", "/rest/v1/item_movement_rows", service=True,
        params={"select": "id,report_id,report_name,report_name_norm,boxes_sold,loose_sold,item_id", "id": f"eq.{row_id}", "limit": "1"},
    )
    if not movement_rows:
        raise HTTPException(404, "سطر الحركة غير موجود")
    movement = movement_rows[0]
    report = await _report(str(movement["report_id"]), profile)
    catalog = await sb(
        "GET", "/rest/v1/item_catalog", service=True,
        params={"select": "id,item_code,item_name,units_per_box", "id": f"eq.{data.item_id}", "limit": "1"},
    )
    if not catalog:
        raise HTTPException(422, "الصنف المختار غير موجود في دليل الأصناف")
    item = catalog[0]
    return await _link_movement_row(movement, report, item, profile)


@router.post("/rows/{row_id}/add-to-catalog")
async def add_report_row_to_catalog(
    row_id: str,
    data: MovementCatalogCreateInput,
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    movement_rows = await sb(
        "GET", "/rest/v1/item_movement_rows", service=True,
        params={
            "select": "id,report_id,report_name,report_name_norm,boxes_sold,loose_sold,item_id",
            "id": f"eq.{row_id}", "limit": "1",
        },
    )
    if not movement_rows:
        raise HTTPException(404, "سطر الحركة غير موجود")
    movement = movement_rows[0]
    if movement.get("item_id"):
        raise HTTPException(409, "هذا الصنف مرتبط بالفعل بدليل الأصناف")
    report = await _report(str(movement["report_id"]), profile)

    code = (data.item_code or "").strip()
    if not code:
        code = f"MOV-{uuid4().hex[:16].upper()}"
    existing = await sb(
        "GET", "/rest/v1/item_catalog", service=True,
        params={"select": "id", "item_code": f"eq.{code}", "limit": "1"},
    )
    if existing:
        raise HTTPException(409, "يوجد صنف بنفس الكود؛ استخدم كودًا آخر أو طابق الصنف الموجود")

    item_name = (data.item_name or movement["report_name"] or "").strip()
    if not item_name:
        raise HTTPException(422, "اسم الصنف مطلوب")
    created = await sb(
        "POST", "/rest/v1/item_catalog", service=True,
        headers={"Prefer": "return=representation"},
        json={
            "item_code": code,
            "item_name": item_name,
            "package_form": (data.package_form or "").strip() or None,
            "units_per_box": data.units_per_box,
        },
    )
    item = created[0]
    linked = await _link_movement_row(movement, report, item, profile)
    return {"ok": True, "item": item, "row": linked, "generated_code": not bool((data.item_code or "").strip())}


@router.delete("/reports/{report_id}")
async def delete_report(report_id: str, profile: dict[str, Any] = Depends(current_profile)) -> Any:
    require_permission(profile, "manage_item_catalog")
    await _report(report_id, profile)
    await sb("DELETE", "/rest/v1/item_movement_reports", service=True, params={"id": f"eq.{report_id}"})
    return {"ok": True}
