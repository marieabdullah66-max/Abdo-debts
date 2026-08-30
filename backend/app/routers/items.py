from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any
import unicodedata

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel

from ..core import (
    SUPABASE_URL, ItemInput, api_headers, current_profile, get_http_client,
    require_permission, sb,
)
from ..xls_biff import read_first_sheet_rows

router = APIRouter(prefix="/api/items", tags=["items"])

MAX_EXCEL_BYTES = 12 * 1024 * 1024
MAX_PREVIEW_ROWS = 20
MAX_PREVIEW_COLS = 30
MAX_IMPORT_ROWS = 50000
XLS_SHEET_NAME = "الورقة الأولى"


class CatalogResetInput(BaseModel):
    confirmation: str


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()




def _normalize_item_name(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "").casefold().strip()
    return " ".join(value.split())


async def _all_catalog_items() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(0, 30000, 1000):
        page = await sb(
            "GET", "/rest/v1/item_catalog", service=True,
            params={
                "select": "id,item_code,item_name,package_form,units_per_box",
                "order": "item_code.asc",
            },
            headers={"Range": f"{offset}-{offset + 999}"},
        )
        rows.extend(page or [])
        if len(page or []) < 1000:
            break
    return rows


async def _all_item_aliases() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for offset in range(0, 50000, 1000):
        page = await sb(
            "GET", "/rest/v1/item_name_aliases", service=True,
            params={"select": "id,report_name,report_name_norm,item_id", "order": "created_at.asc"},
            headers={"Range": f"{offset}-{offset + 999}"},
        )
        rows.extend(page or [])
        if len(page or []) < 1000:
            break
    return rows

def _parse_units(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number <= 0 or int(number) != number:
        return None
    return int(number)


async def _read_excel_bytes(file: UploadFile) -> bytes:
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xls")):
        raise HTTPException(422, "ارفع دليل الأصناف بصيغة .xls أو .xlsx")
    content = await file.read(MAX_EXCEL_BYTES + 1)
    if len(content) > MAX_EXCEL_BYTES:
        raise HTTPException(413, "حجم ملف Excel أكبر من 12 MB")
    if not content:
        raise HTTPException(422, "ملف Excel فارغ")
    return content


def _xls_rows(content: bytes, *, max_rows: int) -> list[list[Any]]:
    try:
        return read_first_sheet_rows(content, max_rows=max_rows, max_cols=MAX_PREVIEW_COLS)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف .xls") from exc


def _content_range_total(value: str | None) -> int | None:
    if not value or "/" not in value:
        return None
    tail = value.rsplit("/", 1)[-1].strip()
    if tail == "*":
        return None
    try:
        return int(tail)
    except ValueError:
        return None


@router.get("")
async def list_items(
    search: str = "",
    limit: int = 100,
    with_meta: bool = False,
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "view_item_analysis")
    max_rows = min(max(limit, 1), 30000)
    params: dict[str, str] = {
        "select": "id,item_code,item_name,package_form,units_per_box,created_at,updated_at",
        "order": "item_name.asc",
    }
    q = search.strip()
    if q:
        safe_q = q
        for ch in "%*(),":
            safe_q = safe_q.replace(ch, " ")
        safe_q = " ".join(safe_q.split())[:120]
        if safe_q:
            params["or"] = f"(item_code.ilike.*{safe_q}*,item_name.ilike.*{safe_q}*)"

    # The first page also asks PostgREST for an exact count. This lets the
    # frontend render only 100 rows by default without downloading the full
    # catalog just to know how many items exist.
    first_size = min(max_rows, 1000)
    headers = api_headers(service=True)
    headers.update({"Range": f"0-{first_size - 1}", "Prefer": "count=exact"})
    response = await get_http_client().request(
        "GET", f"{SUPABASE_URL}/rest/v1/item_catalog", headers=headers, params=params
    )
    if response.status_code >= 400:
        detail: Any = response.text
        try:
            payload = response.json()
            detail = payload.get("message") or payload.get("error_description") or payload.get("msg") or detail
        except Exception:
            pass
        raise HTTPException(response.status_code, detail)

    result: list[dict[str, Any]] = response.json() if response.content else []
    total = _content_range_total(response.headers.get("content-range"))

    # "Show all" is explicit, so only then do we page through the rest.
    page_size = 1000
    for offset in range(first_size, max_rows, page_size):
        last = min(offset + page_size - 1, max_rows - 1)
        rows = await sb(
            "GET", "/rest/v1/item_catalog", service=True, params=params,
            headers={"Range": f"{offset}-{last}"},
        )
        result.extend(rows or [])
        if len(rows or []) < (last - offset + 1):
            break

    if with_meta:
        return {"items": result, "total": total if total is not None else len(result)}
    return result


@router.post("")
async def create_item(data: ItemInput, profile: dict[str, Any] = Depends(current_profile)) -> Any:
    require_permission(profile, "manage_item_catalog")
    code = data.item_code.strip()
    existing = await sb(
        "GET", "/rest/v1/item_catalog", service=True,
        params={"select": "id", "item_code": f"eq.{code}", "limit": "1"},
    )
    if existing:
        raise HTTPException(409, "يوجد صنف بنفس الكود")
    rows = await sb(
        "POST", "/rest/v1/item_catalog", service=True,
        headers={"Prefer": "return=representation"},
        json={
            "item_code": code,
            "item_name": data.item_name.strip(),
            "package_form": (data.package_form or "").strip() or None,
            "units_per_box": data.units_per_box,
        },
    )
    return rows[0]


@router.put("/{item_id}")
async def update_item(item_id: str, data: ItemInput, profile: dict[str, Any] = Depends(current_profile)) -> Any:
    require_permission(profile, "manage_item_catalog")
    code = data.item_code.strip()
    existing = await sb(
        "GET", "/rest/v1/item_catalog", service=True,
        params={"select": "id", "item_code": f"eq.{code}", "id": f"neq.{item_id}", "limit": "1"},
    )
    if existing:
        raise HTTPException(409, "يوجد صنف آخر بنفس الكود")
    rows = await sb(
        "PATCH", "/rest/v1/item_catalog", service=True,
        headers={"Prefer": "return=representation"},
        params={"id": f"eq.{item_id}"},
        json={
            "item_code": code,
            "item_name": data.item_name.strip(),
            "package_form": (data.package_form or "").strip() or None,
            "units_per_box": data.units_per_box,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        },
    )
    if not rows:
        raise HTTPException(404, "الصنف غير موجود")
    return rows[0]


@router.post("/catalog/reset")
async def reset_catalog(
    data: CatalogResetInput,
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    if data.confirmation.strip() != "حذف الدليل":
        raise HTTPException(422, "اكتب: حذف الدليل — لتأكيد العملية")
    result = await sb("POST", "/rest/v1/rpc/reset_item_catalog", service=True, json={})
    return result or {"ok": True}


@router.delete("/{item_id}")
async def delete_item(item_id: str, profile: dict[str, Any] = Depends(current_profile)) -> Any:
    require_permission(profile, "manage_item_catalog")
    await sb("DELETE", "/rest/v1/item_catalog", service=True, params={"id": f"eq.{item_id}"})
    return {"ok": True}


@router.post("/import/preview")
async def preview_item_excel(
    file: UploadFile = File(...),
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    content = await _read_excel_bytes(file)
    filename = (file.filename or "").lower()
    if filename.endswith(".xls"):
        source_rows = _xls_rows(content, max_rows=MAX_PREVIEW_ROWS)
        rows: list[list[str]] = []
        for row in source_rows[:MAX_PREVIEW_ROWS]:
            values = [_clean_text(v) for v in row[:MAX_PREVIEW_COLS]]
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        suggested_columns = None
        if rows and len(rows[0]) >= 5 and _parse_units(rows[0][0]) is not None and _clean_text(rows[0][1]) and _clean_text(rows[0][4]):
            suggested_columns = {"units": 1, "name": 2, "package": 3, "code": 5}
        return {
            "filename": file.filename,
            "sheets": [{"name": XLS_SHEET_NAME, "rows": rows}],
            "suggested_header_row": 0 if suggested_columns else 1,
            "suggested_columns": suggested_columns,
        }

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(422, "تعذر قراءة ملف Excel") from exc

    sheets = []
    for ws in wb.worksheets[:10]:
        rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=MAX_PREVIEW_ROWS, values_only=True):
            values = [_clean_text(v) for v in row[:MAX_PREVIEW_COLS]]
            while values and values[-1] == "":
                values.pop()
            rows.append(values)
        sheets.append({"name": ws.title, "rows": rows})
    wb.close()
    return {"filename": file.filename, "sheets": sheets}


@router.post("/import")
async def import_item_excel(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    header_row: int = Form(...),
    code_column: int = Form(...),
    name_column: int = Form(...),
    units_column: int = Form(...),
    package_column: int | None = Form(None),
    profile: dict[str, Any] = Depends(current_profile),
) -> Any:
    require_permission(profile, "manage_item_catalog")
    if header_row < 0 or min(code_column, name_column, units_column) < 1:
        raise HTTPException(422, "تحديد الأعمدة غير صالح")

    content = await _read_excel_bytes(file)
    filename = (file.filename or "").lower()
    wb = None
    if filename.endswith(".xls"):
        if sheet_name != XLS_SHEET_NAME:
            raise HTTPException(422, "ورقة ملف .xls المحددة غير موجودة")
        all_rows = _xls_rows(content, max_rows=MAX_IMPORT_ROWS)
        row_iter = iter(all_rows[header_row:])
    else:
        try:
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(422, "تعذر قراءة ملف Excel") from exc
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(422, "ورقة Excel المحددة غير موجودة")
        ws = wb[sheet_name]
        row_iter = ws.iter_rows(min_row=header_row + 1, values_only=True)

    # One catalog row per item code. If the same code appears with more than one
    # spelling in the imported file, the last spelling becomes the visible name
    # and all other spellings are learned as aliases for the same catalog item.
    items_by_code: dict[str, dict[str, Any]] = {}
    names_by_code: dict[str, dict[str, str]] = {}
    skipped = 0
    invalid_examples: list[str] = []
    max_col = max(code_column, name_column, units_column, package_column or 0)

    try:
        for row_number, row in enumerate(row_iter, start=header_row + 1):
            values = list(row)
            if not any(v not in (None, "") for v in values):
                continue
            if len(values) < max_col:
                values.extend([None] * (max_col - len(values)))
            code = _clean_text(values[code_column - 1])
            name = _clean_text(values[name_column - 1])
            units = _parse_units(values[units_column - 1])
            package = _clean_text(values[package_column - 1]) if package_column else ""
            if not code or not name or units is None:
                skipped += 1
                if len(invalid_examples) < 8:
                    invalid_examples.append(f"صف {row_number}")
                continue

            code = code[:160]
            name = name[:240]
            normalized_name = _normalize_item_name(name)
            if not normalized_name:
                skipped += 1
                if len(invalid_examples) < 8:
                    invalid_examples.append(f"صف {row_number}")
                continue

            names_by_code.setdefault(code, {})[normalized_name] = name
            # Preserve the previous import behavior: the last row for a repeated
            # code is the current/master representation of that code.
            items_by_code[code] = {
                "item_code": code,
                "item_name": name,
                "package_form": package[:120] or None,
                "units_per_box": units,
            }
    finally:
        if wb is not None:
            wb.close()

    rows = list(items_by_code.values())
    if not rows:
        raise HTTPException(422, "لم يتم العثور على أصناف صالحة للاستيراد حسب الأعمدة المحددة")

    before_catalog = await _all_catalog_items()
    before_by_code = {str(x.get("item_code") or ""): x for x in before_catalog}

    new_items = 0
    same_name_codes = 0
    renamed_codes = 0
    alias_candidates_by_code: dict[str, dict[str, str]] = {}

    for code, incoming in items_by_code.items():
        existing = before_by_code.get(code)
        incoming_norm = _normalize_item_name(str(incoming.get("item_name") or ""))
        candidate_names = dict(names_by_code.get(code) or {})
        if existing:
            old_name = str(existing.get("item_name") or "").strip()
            old_norm = _normalize_item_name(old_name)
            if old_norm == incoming_norm:
                same_name_codes += 1
            else:
                renamed_codes += 1
                if old_norm:
                    candidate_names[old_norm] = old_name
        else:
            new_items += 1

        # The current imported spelling becomes item_catalog.item_name after the
        # upsert, so it does not need a duplicate alias row.
        candidate_names.pop(incoming_norm, None)
        if candidate_names:
            alias_candidates_by_code[code] = candidate_names

    now = datetime.now(timezone.utc).isoformat()
    payload = []
    for item in rows:
        payload.append({**item, "updated_at": now})

    # item_code is the immutable identity: existing codes are updated in place,
    # while unseen codes become new catalog items.
    for start in range(0, len(payload), 400):
        await sb(
            "POST", "/rest/v1/item_catalog?on_conflict=item_code", service=True,
            headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
            json=payload[start:start + 400],
        )

    after_catalog = await _all_catalog_items()
    after_by_code = {str(x.get("item_code") or ""): x for x in after_catalog}
    canonical_norm_to_ids: dict[str, set[str]] = {}
    for item in after_catalog:
        norm = _normalize_item_name(str(item.get("item_name") or ""))
        if norm:
            canonical_norm_to_ids.setdefault(norm, set()).add(str(item.get("id") or ""))

    existing_alias_rows = await _all_item_aliases()
    existing_aliases = {
        str(x.get("report_name_norm") or ""): str(x.get("item_id") or "")
        for x in existing_alias_rows
        if x.get("report_name_norm")
    }

    pending_aliases: dict[str, dict[str, Any]] = {}
    blocked_alias_norms: set[str] = set()
    aliases_already_known = 0
    alias_conflicts = 0
    alias_conflict_examples: list[str] = []

    for code, names in alias_candidates_by_code.items():
        item = after_by_code.get(code)
        if not item:
            continue
        item_id = str(item.get("id") or "")
        for norm, display_name in names.items():
            if not norm or norm in blocked_alias_norms:
                continue

            alias_item_id = existing_aliases.get(norm)
            if alias_item_id:
                if alias_item_id == item_id:
                    aliases_already_known += 1
                else:
                    alias_conflicts += 1
                    if len(alias_conflict_examples) < 8:
                        alias_conflict_examples.append(f"{display_name} — الكود {code}")
                continue

            # A name that is already the canonical name of another code is
            # ambiguous. Do not silently teach a wrong mapping.
            canonical_ids = canonical_norm_to_ids.get(norm) or set()
            if canonical_ids and canonical_ids != {item_id}:
                alias_conflicts += 1
                if len(alias_conflict_examples) < 8:
                    alias_conflict_examples.append(f"{display_name} — الكود {code}")
                continue
            if canonical_ids == {item_id}:
                aliases_already_known += 1
                continue

            previous_pending = pending_aliases.get(norm)
            if previous_pending:
                if str(previous_pending.get("item_id") or "") == item_id:
                    aliases_already_known += 1
                else:
                    # The same spelling appeared as an alias for two different
                    # codes in this import. Do not pick one silently.
                    alias_conflicts += 1
                    pending_aliases.pop(norm, None)
                    blocked_alias_norms.add(norm)
                    if len(alias_conflict_examples) < 8:
                        alias_conflict_examples.append(f"{display_name} — الكود {code}")
                continue

            pending_aliases[norm] = {
                "report_name": display_name[:300],
                "report_name_norm": norm[:300],
                "item_id": item_id,
                "created_by": profile.get("id"),
                "updated_at": now,
            }

    aliases_to_insert = list(pending_aliases.values())
    for start in range(0, len(aliases_to_insert), 400):
        await sb(
            "POST", "/rest/v1/item_name_aliases?on_conflict=report_name_norm", service=True,
            headers={"Prefer": "resolution=ignore-duplicates,return=minimal"},
            json=aliases_to_insert[start:start + 400],
        )

    return {
        "ok": True,
        "imported": len(rows),
        "processed_codes": len(rows),
        "new_items": new_items,
        "existing_codes": len(rows) - new_items,
        "same_name_codes": same_name_codes,
        "renamed_codes": renamed_codes,
        "aliases_added": len(aliases_to_insert),
        "aliases_already_known": aliases_already_known,
        "alias_conflicts": alias_conflicts,
        "alias_conflict_examples": alias_conflict_examples,
        "skipped": skipped,
        "invalid_examples": invalid_examples,
        "message": (
            "تم تحديث دليل الأصناف حسب الكود؛ الاسم الأحدث أصبح الاسم الظاهر، "
            "والأسماء السابقة/البديلة تم حفظها لنفس الصنف حتى تتعرف عليها تقارير الحركة"
        ),
    }

